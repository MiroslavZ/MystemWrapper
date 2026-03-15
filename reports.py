from datetime import datetime
from logging import getLogger
from pathlib import Path

from openpyxl import Workbook

logger = getLogger(__name__)


def save_report(result_vectors: list[dict[tuple[str, str], int]], original_files: list[Path]) -> None:
    logger.info('Создание отчета')
    workbook = Workbook()
    worksheet = workbook.active
    current_date = datetime.now()

    # Шаг отступа для каждого нового файла.
    # 4 столбца данных (Слово, Часть речи, Частота, Доля) + 1 пустой столбец-разделитель
    COLUMNS_PER_FILE = 5

    for index, vector in enumerate(result_vectors):
        # Вычисляем номер начального столбца для текущего файла (1, 6, 11, ...)
        base_col = 1 + (index * COLUMNS_PER_FILE)

        # Текущая строка для записи (сбрасывается в 1 для каждого нового файла)
        current_row = 1

        # 1. Записываем имя файла (заголовок блока)
        worksheet.cell(row=current_row, column=base_col, value=original_files[index].stem)
        current_row += 1

        # 2. Заголовки столбцов
        headers = ['Слово', 'Часть речи', 'Частота', 'Доля']
        for i, header in enumerate(headers):
            worksheet.cell(row=current_row, column=base_col + i, value=header)

        current_row += 1

        # Подготовка к циклу данных
        freq_sum = sum(vector.values())
        if freq_sum == 0:
            freq_sum = 1  # Защита от деления на ноль

        data_start_row = current_row

        # 3. Запись данных построчно
        for (word, pos), freq in vector.items():
            share = round(freq / freq_sum, 6)

            # Пишем в соответствующие столбцы относительно base_col
            worksheet.cell(row=current_row, column=base_col, value=word)  # Слово
            worksheet.cell(row=current_row, column=base_col + 1, value=pos)  # Часть речи
            worksheet.cell(row=current_row, column=base_col + 2, value=freq)  # Частота
            worksheet.cell(row=current_row, column=base_col + 3, value=share)  # Доля

            current_row += 1

        data_end_row = current_row - 1

        # 4. Запись метаданных (координат), если были данные
        if data_end_row >= data_start_row:
            # Координаты столбцов с Частотой и Долей
            freq_col_idx = base_col + 2
            share_col_idx = base_col + 3

            # Получаем строковые координаты (например, "C3", "C100")
            freq_start_cell = worksheet.cell(row=data_start_row, column=freq_col_idx).coordinate
            freq_end_cell = worksheet.cell(row=data_end_row, column=freq_col_idx).coordinate

            share_start_cell = worksheet.cell(row=data_start_row, column=share_col_idx).coordinate
            share_end_cell = worksheet.cell(row=data_end_row, column=share_col_idx).coordinate

            # Пишем служебную информацию под таблицей данных
            # Строка "Частоты", start, end
            worksheet.cell(row=current_row, column=base_col, value='Частоты')
            worksheet.cell(row=current_row, column=base_col + 1, value=freq_start_cell)
            worksheet.cell(row=current_row, column=base_col + 2, value=freq_end_cell)
            current_row += 1

            # Строка "Доли", start, end
            worksheet.cell(row=current_row, column=base_col, value='Доли')
            worksheet.cell(row=current_row, column=base_col + 1, value=share_start_cell)
            worksheet.cell(row=current_row, column=base_col + 2, value=share_end_cell)

    filename = f'Statistics {current_date.strftime("%Y-%m-%d %H %M %S")}.xlsx'
    workbook.save(filename)
    path = Path(f'./{filename}').resolve()
    logger.info(f'Файл сохранен {path}')